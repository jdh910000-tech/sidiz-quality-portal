# -*- coding: utf-8 -*-
"""
하자유형(표준화) 1차 클렌징 안 생성기
- 입력: 클레임현황_업로드일_전체제품_2026-04-01~2026-04-30_20260430.xlsx
- 출력: cleanse_proposal_v1.xlsx (사용자 검토용)

규칙:
  (1) 띄어쓰기/공백/구두점만 다른 항목은 자동 머지 (제일 자주 쓰인 표기로 통일)
  (2) 괄호 안 노이즈(연수/기간/날짜)는 제거.  부품/원인/맥락은 유지
       - 노이즈 정규식:
         · ^\d+년( \d+개월)?$        예: '3년', '24년 8월', '2년 6개월'
         · ^\d+년( \d+월)?$          예: '26년 1월'
         · ^\d+개월$
         · ^\d+\.\d+$               예: '24.5'
       - 그 외 텍스트(중심봉·락커·팔걸이·리머추정·설변전·좌측 등)는 유지
  (3) 콤마로 분리된 괄호 내용은 토큰별로 (2)를 적용 후 의미있는 토큰만 남김
"""
import openpyxl, re, sys
from collections import Counter, defaultdict
from openpyxl.styles import PatternFill, Font, Alignment

INPUT  = r'C:\Users\FURSYS\Downloads\클레임현황_업로드일_전체제품_2026-04-01~2026-04-30_20260430.xlsx'
OUTPUT = r'C:\Users\FURSYS\Desktop\클로드\cleanse_proposal_v1.xlsx'

NOISE_PATTERNS = [
    re.compile(r'^\d+\s*년(\s*\d+\s*개?월)?$'),       # 3년, 24년 8월, 2년 6개월
    re.compile(r'^\d+\s*년(\s*\d+\s*월)?$'),          # 26년 1월
    re.compile(r'^\d+\s*개월$'),                       # 6개월
    re.compile(r'^\d+\.\d+년?$'),                      # 24.5
    re.compile(r'^\d{4}[-./]\d{1,2}([-./]\d{1,2})?$'), # 2024-01, 2024.01.05
]

def is_noise_token(t):
    t = t.strip()
    if not t: return True
    for p in NOISE_PATTERNS:
        if p.match(t): return True
    return False

def clean_paren(content):
    """괄호 안 콘텐츠를 받아 노이즈 토큰을 제거한 결과 반환. 빈 결과면 None"""
    # ',' 또는 '/' 또는 '·' 로 토큰 분리
    parts = re.split(r'[,/·]', content)
    keep = [p.strip() for p in parts if not is_noise_token(p)]
    if not keep:
        return None
    return ', '.join(keep)

def clean_label(label):
    """라벨에서 괄호 노이즈 제거 + 공백 정규화"""
    s = str(label).strip()
    # 괄호 처리: ( ... ) 또는 （...）
    def repl(m):
        inner = m.group(1)
        cleaned = clean_paren(inner)
        if cleaned is None:
            return ''
        return f'({cleaned})'
    s = re.sub(r'\(([^()]*)\)', repl, s)
    s = re.sub(r'（([^（）]*)）', repl, s)
    # 공백 정규화: 다중 공백 → 단일 공백
    s = re.sub(r'\s+', ' ', s).strip()
    # '괄호 앞 공백' 일관화: ' (' → '('
    s = re.sub(r'\s+\(', '(', s)
    # 끝부분 ',' 등 정리
    s = re.sub(r'[,\s]+$', '', s)
    return s

def norm_key(s):
    """비교용 정규화: 공백 제거 + 소문자 + 흔한 동의어"""
    s = re.sub(r'\s+', '', s.lower())
    return s

def main():
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    ws = wb['클레임_업로드일기준']
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    H = {h:i for i,h in enumerate(headers)}

    rows = []
    for r in range(2, ws.max_row+1):
        rows.append([ws.cell(r,c+1).value for c in range(len(headers))])

    # 라벨별 빈도
    label_count = Counter()
    label_examples = defaultdict(list)  # label -> [(item, claim_detail_excerpt)]
    label_category = defaultdict(Counter)  # label -> Counter(category)
    for r in rows:
        v = r[H['하자유형(표준화)']]
        if not v: continue
        k = str(v).strip()
        label_count[k] += 1
        if len(label_examples[k]) < 2:
            cd = r[H['요구내역']] or ''
            cd = re.sub(r'\s+', ' ', str(cd))[:80]
            label_examples[k].append((r[H['제품']] or '-', cd))
        cat = r[H['defect_category']] or '미분류'
        label_category[k][cat] += 1

    print(f'unique labels: {len(label_count)}')

    # 1단계: 클렌징 후 라벨 → 정규화 키로 묶기
    rows_out = []
    cleaned_to_originals = defaultdict(list)
    for label, n in label_count.most_common():
        cleaned = clean_label(label)
        cleaned_to_originals[norm_key(cleaned)].append((label, cleaned, n))

    # 2단계: 같은 정규화 키 그룹 안에서 대표 라벨 선택 (가장 빈도 높은 cleaned 형태)
    review_rows = []
    auto_merge_count = 0
    paren_strip_count = 0
    no_change_count = 0
    for nk, items in cleaned_to_originals.items():
        # 대표 cleaned 형태 선정: 빈도 합이 가장 큰 cleaned 표기
        cleaned_count = Counter()
        for _orig, _clean, n in items:
            cleaned_count[_clean] += n
        canonical = cleaned_count.most_common(1)[0][0]

        # 그룹 내 처리 분류 표시용
        for orig, clean, n in items:
            if orig == canonical:
                action = '유지'
                no_change_count += 1
            elif clean != orig and norm_key(orig) == norm_key(clean):
                # 클렌징도 안 됐는데 다른 표기와 정규화 후 같으면 단순 띄어쓰기 차이
                action = '자동머지(공백차이)'
                auto_merge_count += 1
            elif clean != orig:
                action = '괄호노이즈제거'
                paren_strip_count += 1
            else:
                action = '자동머지(중복)'
                auto_merge_count += 1

            top_cat = label_category[orig].most_common(1)[0][0] if label_category[orig] else '-'
            ex_item, ex_detail = label_examples[orig][0] if label_examples[orig] else ('','')
            review_rows.append({
                '원본 라벨': orig,
                '건수': n,
                '제안 표준 라벨': canonical,
                '클렌징 결과': clean,
                '처리분류': action,
                '주카테고리': top_cat,
                '예시 제품': ex_item,
                '예시 요구내역(80자)': ex_detail,
            })

    # 정렬: 그룹별로 묶이게 (canonical 기준), 그 안은 빈도 desc
    review_rows.sort(key=lambda r: (r['제안 표준 라벨'], -r['건수']))

    # 출력 엑셀
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '클렌징_검토'

    header_cols = ['원본 라벨','건수','제안 표준 라벨','클렌징 결과','처리분류','주카테고리','예시 제품','예시 요구내역(80자)','확정(O/X/수정)']
    for i, h in enumerate(header_cols, 1):
        c = ws_out.cell(1, i, h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='002BD2')
        c.alignment = Alignment(vertical='center', horizontal='center')

    fill_merge = PatternFill('solid', fgColor='FFF3D6')   # 자동머지
    fill_paren = PatternFill('solid', fgColor='E8F4FF')   # 괄호제거
    fill_keep  = PatternFill('solid', fgColor='F0F0F5')   # 유지

    prev_canon = None
    row_idx = 2
    for r in review_rows:
        # 그룹 경계 표시 (canonical 바뀔 때 굵은 위 테두리)
        is_group_first = (prev_canon != r['제안 표준 라벨'])
        prev_canon = r['제안 표준 라벨']
        for i, key in enumerate(['원본 라벨','건수','제안 표준 라벨','클렌징 결과','처리분류','주카테고리','예시 제품','예시 요구내역(80자)'], 1):
            c = ws_out.cell(row_idx, i, r[key])
            if r['처리분류'].startswith('자동머지'):
                c.fill = fill_merge
            elif r['처리분류'] == '괄호노이즈제거':
                c.fill = fill_paren
            else:
                c.fill = fill_keep
        # 9번째 열: 확정 칸
        ws_out.cell(row_idx, 9, '')
        row_idx += 1

    # 컬럼 너비
    widths = [25, 8, 25, 25, 18, 12, 12, 50, 14]
    for i, w in enumerate(widths, 1):
        ws_out.column_dimensions[chr(64+i)].width = w

    # 요약 시트
    ws_sum = wb_out.create_sheet('요약')
    canon_count = Counter()
    for r in review_rows:
        canon_count[r['제안 표준 라벨']] += r['건수']
    summary = [
        ('전체 원본 라벨 수', len(review_rows)),
        ('제안 표준 라벨 수', len(canon_count)),
        ('축소율', f'{(1 - len(canon_count)/len(review_rows))*100:.1f}%'),
        ('처리분류: 유지', no_change_count),
        ('처리분류: 자동머지', auto_merge_count),
        ('처리분류: 괄호노이즈제거', paren_strip_count),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ws_sum.cell(i, 1, k).font = Font(bold=True)
        ws_sum.cell(i, 2, v)
    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 20

    # 표준 라벨별 통합 결과 시트
    ws_canon = wb_out.create_sheet('표준라벨_통합결과')
    ws_canon.cell(1,1,'제안 표준 라벨').font = Font(bold=True, color='FFFFFF')
    ws_canon.cell(1,1).fill = PatternFill('solid', fgColor='002BD2')
    ws_canon.cell(1,2,'합계 건수').font = Font(bold=True, color='FFFFFF')
    ws_canon.cell(1,2).fill = PatternFill('solid', fgColor='002BD2')
    ws_canon.cell(1,3,'통합된 원본 표기').font = Font(bold=True, color='FFFFFF')
    ws_canon.cell(1,3).fill = PatternFill('solid', fgColor='002BD2')

    canon_orig_map = defaultdict(list)
    for r in review_rows:
        canon_orig_map[r['제안 표준 라벨']].append((r['원본 라벨'], r['건수']))

    rr = 2
    for canon, total in canon_count.most_common():
        origs = canon_orig_map[canon]
        ws_canon.cell(rr, 1, canon)
        ws_canon.cell(rr, 2, total)
        if len(origs) == 1 and origs[0][0] == canon:
            ws_canon.cell(rr, 3, '— (변경 없음)')
        else:
            ws_canon.cell(rr, 3, ' / '.join(f'{o}({n})' for o,n in origs))
        rr += 1
    ws_canon.column_dimensions['A'].width = 30
    ws_canon.column_dimensions['B'].width = 12
    ws_canon.column_dimensions['C'].width = 80

    # ─── 추가 검토 후보 시트 (편집거리 기반 잠재 머지) ───
    ws_fuzzy = wb_out.create_sheet('추가검토_후보')
    fuzzy_headers = ['라벨 A', '건수 A', '라벨 B', '건수 B', '편집거리', '추정 사유', '확정(O/X/수정)']
    for i, h in enumerate(fuzzy_headers, 1):
        c = ws_fuzzy.cell(1, i, h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='8B4513')

    def edit_distance(a, b):
        if a == b: return 0
        if abs(len(a)-len(b)) > 3: return 99
        if len(a) > len(b): a,b = b,a
        prev = list(range(len(a)+1))
        for j, cb in enumerate(b, 1):
            curr = [j] + [0]*len(a)
            for i, ca in enumerate(a, 1):
                curr[i] = min(prev[i]+1, curr[i-1]+1, prev[i-1] + (0 if ca==cb else 1))
            prev = curr
        return prev[-1]

    canons = list(canon_count.keys())
    fuzzy_rows = []
    seen = set()
    for i, a in enumerate(canons):
        for b in canons[i+1:]:
            # 너무 짧거나 차이가 너무 크면 스킵
            if len(a) < 3 or len(b) < 3: continue
            if abs(len(a) - len(b)) > 2: continue
            d = edit_distance(a, b)
            # 짧은 라벨일수록 임계 낮춤
            min_len = min(len(a), len(b))
            threshold = 1 if min_len <= 6 else 2
            if 0 < d <= threshold:
                pair = tuple(sorted([a,b]))
                if pair in seen: continue
                seen.add(pair)
                # 사유 추정
                reason = '오타/공백/유사표현 추정'
                if a.replace(' ','') == b.replace(' ',''):
                    reason = '공백 차이 (이미 처리되었어야 함)'
                fuzzy_rows.append((a, canon_count[a], b, canon_count[b], d, reason))

    fuzzy_rows.sort(key=lambda x: (x[4], -(x[1]+x[3])))
    for ridx, (a, na, b, nb, d, reason) in enumerate(fuzzy_rows, 2):
        ws_fuzzy.cell(ridx, 1, a)
        ws_fuzzy.cell(ridx, 2, na)
        ws_fuzzy.cell(ridx, 3, b)
        ws_fuzzy.cell(ridx, 4, nb)
        ws_fuzzy.cell(ridx, 5, d)
        ws_fuzzy.cell(ridx, 6, reason)
        # 음영
        fill = PatternFill('solid', fgColor='FFE4B5' if d == 1 else 'FFF8DC')
        for ci in range(1, 7):
            ws_fuzzy.cell(ridx, ci).fill = fill

    for col, w in zip('ABCDEFG', [25, 8, 25, 8, 10, 25, 14]):
        ws_fuzzy.column_dimensions[col].width = w

    print(f'fuzzy candidates: {len(fuzzy_rows)}')

    wb_out.save(OUTPUT)
    print(f'saved: {OUTPUT}')
    print(f'summary: {summary}')

if __name__ == '__main__':
    sys.stdout.reconfigure(encoding='utf-8')
    main()
