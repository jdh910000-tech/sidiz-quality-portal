import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\FURSYS\Downloads\제품별클레임_전체제품_2026-01-01~2026-06-18_20260618.xlsx',
    data_only=True
)

print('=== 시트 목록 ===')
for i, name in enumerate(wb.sheetnames):
    print(f'  {i}: {name}')

print()
ws = wb.worksheets[1]
print(f'=== {ws.title} (rows={ws.max_row}) ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
    vals = [str(v) if v is not None else '' for v in row]
    if any(v.strip() for v in vals):
        print(f'R{i+1:02d}: ' + ' | '.join(vals))
